import sys
import os
import re
import timeit
import time
from time import clock
import datetime

import numpy as np

import pandas as pd
from pandas import options

from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base

import xlrd
from openpyxl import *

from matplotlib import pyplot as plt
import matplotlib
from matplotlib_venn import venn3, venn3_circles,venn2
matplotlib.use('Agg')

options.io.excel.xlsx.writer= 'xlsxwriter'


def download_pandas(name,sheet):
  xls = pd.read_excel('/Users/agatorano/Code/metascape/metascape.org/media/'+name,index_col=None,sheetname=sheet, na_values=['NA'])
  return xls

#def get_column_pandas(name):
#  xls=download_pandas(name)
#  col = xls[xls.columns.tolist()[0]]
#  return col

def get_column_pandas(name,sheet):
  xls=download_pandas(name,sheet)
  col = xls[xls.columns.tolist()[0]]
  col = col.tolist()
  return col

def save_excel(data,writer,index):

  data.to_excel(writer,'Processed_Data_%s'%index,index=False)
  
def add_cols(list_,name,sheet):
   
  synonym=[]
  description=[]
  type_of_gene=[]
  #full_name=[]
  symbol = []
  gid=[]

  col = get_column_pandas(name,sheet)
  id_type = get_id_type(col[0])

  for row in list_:
    gid.append(row[0])
    symbol.append(row[2])
    synonym.append(row[3])
    description.append(row[4])
    type_of_gene.append(row[5])
    #full_name.append(row[6])

  data = download_pandas(name,sheet)
  gid = list(map(str,gid))

  data['gid']=gid
  data['symbol']=symbol
  data['synonym']=synonym
  data['description']=description
  data['type_of_gene']=type_of_gene
  #data['full_name'] = full_name

  return data

def add_annotation(list_):

  engine = create_engine('mysql+mysqlconnector://ag:529382Ag@localhost/gp?charset=utf8&use_unicode=0',pool_recycle=3600)

  #query = "select a.gid as entrez,a.content as synonym,b.content as description,c.content as type_of_gene,d.content as full_name from (select gid,content from annotation where annotation.annotation_type_id=1) a,(select gid,content from annotation where annotation.annotation_type_id=2) b,(select gid,content from annotation where annotation.annotation_type_id=3) c,(select gid,content from annotation where annotation.annotation_type_id=4) d where a.gid=b.gid and b.gid = c.gid and c.gid = d.gid"
  query = "select a.gid as entrez,a.content as synonym,b.content as description,c.content as type_of_gene from (select gid,content from annotation where annotation.annotation_type_id=1) a,(select gid,content from annotation where annotation.annotation_type_id=2) b,(select gid,content from annotation where annotation.annotation_type_id=3) c where a.gid=b.gid and b.gid = c.gid"

  result2 = engine.execute(query)
  result = []

  for row in result2:
    #result.append({"entrez":row['entrez'],"synonym":row['synonym'],"description":row['description'],"type_of_gene":row['type_of_gene'],"full_name":row["full_name"]})
    result.append({"entrez":row['entrez'],"synonym":row['synonym'],"description":row['description'],"type_of_gene":row['type_of_gene']})

  result2.close()


  count = 0
  for group in list_:
    flag=0
    for row in result:
      if  group[0] == row['entrez']:
        list_[count].append(row['synonym'])
        list_[count].append(row['description'])
        list_[count].append(row['type_of_gene'])
        #list_[count].append(row['full_name'])
        flag=1
        break
    if flag ==0:
      list_[count].append('NA')
      list_[count].append('NA')
      list_[count].append('NA')
      #list_[count].append('NA')

    count=count+1

  return list_  

def get_id_type(el):

  entrez = re.compile('^[0-9]+$')
  refseq = re.compile('^NM_')

  if(entrez.match(str(el))):
    return '0'
  
  if(refseq.match(str(el))):
    return '3'
  
  else:
    return '1'

def get_results(id_type):

  ids = [1,3,0]


  id_dict = {0:"gid",1:"symbol",2:"refseq_protein",3:'refseq',4:"refseq_gene"}

  engine = create_engine('mysql+mysqlconnector://ag:529382Ag@localhost/gp?charset=utf8&use_unicode=0',pool_recycle=3600)

  if(int(id_type)!=0):
    query = "select a.gid as entrez, a.source_id as "+str(id_dict[int(id_type)])+",b.source_id as symbol from (select * from gid2source_id where gid2source_id.id_type_id =" +str(id_type)+") a,(select * from gid2source_id where gid2source_id.id_type_id = 1) b where a.gid =b.gid;"
  else:
    query = "select a.gid as entrez, a.gid as "+str(id_dict[int(id_type)])+",b.source_id as symbol from (select * from gid2source_id) a,(select * from gid2source_id where gid2source_id.id_type_id=1) b where a.gid=b.gid;"


  result1 = engine.execute(query)

  result = {}
  for row in result1:
    result[row[id_dict[int(id_type)]]]={"entrez":row['entrez'],"symbol":row['symbol']}

  result1.close()

  return result

def get_homologues():

  engine = create_engine('mysql+mysqlconnector://ag:529382Ag@localhost/gp?charset=utf8&use_unicode=0',pool_recycle=3600)

  query = "select homologene_id as homologene, gid from homologene;"

  result1 = engine.execute(query)

  result = {}
  
  for row in result1:
    result[row['homologene']]=row['gid']

  return result


def get_gid(name,sheet):
  
  
  col = get_column_pandas(name,sheet)
  gene_ids = []

  id_type = get_id_type(col[0])
  id_dict = {0:"entrez",1:"symbol",2:"refseq_protein",3:'refseq',4:"refseq_gene"}

  result = get_results(id_type)
#  homologue = get_homologues()

  for id in col:
    count = 0
    if id in result:
      count=1
      row = result[id]
      gene_ids.append([row['entrez'],id,row['symbol']])
#    elif homologue[id] in result:
#      homol=homologue[id]
#      count=1
#      row = result[homol]
#      gene_ids.append([homol,id,row['symbol']])

    if count ==0:
      gene_ids.append(['NA','NA','NA'])

  return gene_ids

def process_multiple(names):

  writer = pd.ExcelWriter('/Users/agatorano/Code/metascape/metascape.org/media/%s'%names[-1])
  files_=[]
  #print(names[:len(names)-1])
  for n in names[:len(names)-1]:
    files_.append(pd.ExcelFile('/Users/agatorano/Code/metascape/metascape.org/media/%s'%n))

  for f,i in zip(files_,range(len(files_))):
    df = f.parse(f.sheet_names[0])
    df.to_excel(writer,'Sheet%s'%i)
  writer.save()

  xls = xlrd.open_workbook(r'/Users/agatorano/Code/metascape/metascape.org/media/%s'%names[-1], on_demand=True)
  
  name = names[-1]
  data = []
  genes = []
  for sheet in xls.sheet_names():
    list_ = get_gid(name,sheet)
    list_ = add_annotation(list_)
    genes.append(set([x[0] for x in list_]))
    #print(genes)
    data.append(add_cols(list_,name,sheet))

  writer = pd.ExcelWriter('/Users/agatorano/Code/metascape/metascape.org/media/'+name)
  for i in range(len(data)):
    save_excel(data[i],writer,i+1)

  plt.figure(figsize=(7,7))
  now = datetime.datetime.now()
  path = 'img/'+now.strftime("%Y/%m/%d/venn%H_%M_%S.png")
  img = '/Users/agatorano/Code/metascape/metascape.org/media/'+path
  output_directory = os.path.dirname(img)
  if not os.path.exists(output_directory):
      os.makedirs(output_directory)
  if(len(genes)==3):
    venn3(genes, ('File1', 'File2', 'File3'))
    plt.savefig(img)
  elif(len(genes)==2):
    venn2(genes, ('File1', 'File2'))
    plt.savefig(img)

  writer.save()

  return data,path

def process_single(name,out):

  list_ = get_gid(name,0)
  list_ = add_annotation(list_)
  data=add_cols(list_,name,0)

  writer = pd.ExcelWriter('/Users/agatorano/Code/metascape/metascape.org/media/'+out)
  save_excel(data,writer,1)

  return data


def membership_data(name,data):

  xls = xlrd.open_workbook(r'/Users/agatorano/Code/metascape/metascape.org/media/%s'%name, on_demand=True)
  sheet_names = xls.sheet_names()
  sheet_number=len(sheet_names)
  xls = load_workbook(filename='/Users/agatorano/Code/metascape/metascape.org/media/%s'%name)
  membership={}
  count=1
  for sheet in data:
    for el in sheet['gid']:
      if el in membership:
        if count in membership[el]:
          next
        else:
          membership[el][str(count)]=1
      else:
        membership[el]={}
        membership[el][str(count)]=1
    count+=1


  membership = membership_tidy(membership,sheet_number)

  worksheet=xls.create_sheet(title='multiple_summary')
  add_membership_map(membership,worksheet,sheet_names)

  xls.save(filename='/Users/agatorano/Code/metascape/metascape.org/media/%s'%name)
  
  return membership

def membership_tidy(membership,sheets):

  sheets = [str(x+1) for x in range(sheets)]

  for id_ in membership:
    for i in sheets:
      if i not in membership[id_]:
        membership[id_][i]=0
        
  return membership

def add_membership_map(membership,workbook,col_names):

  col_names.insert(0,'gid')
  col_names.append('total')
  for i in range(len(col_names)):
    workbook.cell(row=1,column=i+1).value=col_names[i]

  row=2
  col=1 

  for id_ in membership:
    col=1
    workbook.cell(row=row,column=col).value=id_
    sum=0
    for sheet in membership[id_]:
      col=col+1
      workbook.cell(row=row,column=col).value=membership[id_][sheet]
      sum+=int(membership[id_][sheet])
    col+=1
    workbook.cell(row=row,column=col).value=sum
    row+=1

def main():

  list_ = get_gid(name)
  list_ = add_annotation(list_)
  data = add_cols(list_,name)
  save_excel(data,name)

if __name__=='__main__':
  main()

